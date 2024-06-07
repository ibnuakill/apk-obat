import tkinter as tk
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from PIL import Image, ImageTk
import os
from datetime import date
import openpyxl
from openpyxl import Workbook
import pathlib

# Warna dan Background
background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

# Inisialisasi Aplikasi
root = tk.Tk()
root.title("Aplikasi Stok Obat")
root.geometry("1250x650+210+100")
root.config(bg=background)

# Cek dan buat file Excel jika tidak ada
file = pathlib.Path('data_obat.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet.append(["Registration No.", "Nama Obat", "Golongan", "Brand", "Expired", "Kemasan", "Jumlah", "Kode Obat", "Distributor", "Harga Beli", "Harga Grosir", "Harga Jual"])
    file.save('data_obat.xlsx')

# Variabel
Registration = StringVar()
Date = StringVar()
Name = StringVar()
Brand = StringVar()
Exp = StringVar()
Kemasan = StringVar()
Jumlah = StringVar()
Kode = StringVar()
Distributor = StringVar()
Beli = IntVar()
Grosir = IntVar()
Jual = IntVar()

# Fungsi
def Exit():
    root.destroy()

def showimage():
    global filename
    global photo2
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Pilih File Gambar", filetype=(("JPG File", ".jpg"), ("PNG File", ".png"), ("All Files", ".")))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

def registration_no():
    file = openpyxl.load_workbook('data_obat.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")

def Clear():
    global img
    Name.set('')
    Brand.set('')
    Exp.set('')
    Kemasan.set('')
    Jumlah.set("Pilih jumlah")
    Kode.set('')
    Distributor.set('')
    Beli.set(0)
    Grosir.set(0)
    Jual.set(0)
    registration_no()
    saveButton.config(state='normal')
    img1 = PhotoImage(file='images/upload_photo.png')
    lbl.config(image=img1)
    lbl.image = img1
    img = ""

def Save():
    R1 = Registration.get()
    D1 = Date.get()
    N1 = Name.get()
    S1 = radio.get()
    B1 = Brand.get()
    E1 = Exp.get()
    K1 = Kemasan.get()
    J1 = Jumlah.get()
    C1 = Kode.get()
    D2 = Distributor.get()
    B2 = Beli.get()
    G1 = Grosir.get()
    J2 = Jual.get()

    if N1 == "" or S1 == "" or B1 == "" or E1 == "" or K1 == "" or J1 == "Pilih jumlah" or C1 == "" or D2 == "" or B2 == 0 or G1 == 0 or J2 == 0:
        messagebox.showerror("Error", "Data ada yang kosong")
    else:
        file = openpyxl.load_workbook('data_obat.xlsx')
        sheet = file.active
        sheet.append([R1, N1, S1, B1, E1, K1, J1, C1, D2, B2, G1, J2])
        file.save('data_obat.xlsx')
        try:
            img.save("obat_images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("Info", "Photo is not available!!")

        messagebox.showinfo("Info", "Data berhasil disubmit")

# Frame Atas
Label(root, text="Email : winanf19@gmail.com ", width=10, height=3, fg="#ADFF2F", anchor='e').pack(side=TOP, fill=X)
Label(root, text="Registrasi Obat Apotik", width=10, height=2, bg="#fff05f", fg="#808080", font='arial 20 bold').pack(side=TOP, fill=X)

# Searchbox
search = StringVar()
Entry(root, textvariable=search, width=15, bd=2, font="arial 20").place(x=820, y=70)
imegicon3 = PhotoImage(file="images/search.png")
Srch = Button(root, text="Search", compound=LEFT, image=imegicon3, width=123, fg="#68ddfa", font="arial 13 bold")
Srch.place(x=1060, y=66)

# Tombol update
imageicon4 = PhotoImage(file="images/update.png")
update_button = Button(root, image=imageicon4, bg="#c36464")
update_button.place(x=110, y=64)

# Registrasi dan tanggal
Label(root, text="Registration No:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Tanggal:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)
Date.set(d1)

# Detail barang
obj = LabelFrame(root, text="Detail Barang", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

# Detail harga
obj2 = LabelFrame(root, text="Detail Harga", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj, text="Nama Obat", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Golongan", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Brand", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)
Label(obj, text="Expired", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Kemasan", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Jumlah", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=50)

radio = IntVar()
R1 = Radiobutton(obj, text="Blasa", variable=radio, value=1, bg=framebg, fg=framefg)
R1.place(x=150, y=100)
R2 = Radiobutton(obj, text="Keras", variable=radio, value=2, bg=framebg, fg=framefg)
R2.place(x=210, y=100)

brand_entry = Entry(obj, textvariable=Brand, width=20, font="arial 10")
brand_entry.place(x=160, y=150)

expired_entry = Entry(obj, textvariable=Exp, width=20, font="arial 10")
expired_entry.place(x=630, y=50)

kemasan_entry = Entry(obj, textvariable=Kemasan, width=20, font="arial 10")
kemasan_entry.place(x=630, y=100)

Jumlah = Combobox(obj, values=list(range(1, 1001)), font="rabato 10", width=17, state="r")
Jumlah.place(x=630, y=150)
Jumlah.set("Select Jumlah")

Label(obj2, text="Kode Obat", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Distributor", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj2, text="Harga Beli", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj2, text="Harga Grosir", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Harga Jual", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

kode_obat_entry = Entry(obj2, textvariable=Kode, width=20, font="arial 10")
kode_obat_entry.place(x=160, y=50)

distributor_entry = Entry(obj2, textvariable=Distributor, width=20, font="arial 10")
distributor_entry.place(x=160, y=100)

harga_beli_entry = Entry(obj2, textvariable=Beli, width=20, font="arial 10")
harga_beli_entry.place(x=160, y=150)

harga_grosir_entry = Entry(obj2, textvariable=Grosir, width=20, font="arial 10")
harga_grosir_entry.place(x=630, y=50)

harga_jual_entry = Entry(obj2, textvariable=Jual, width=20, font="arial 10")
harga_jual_entry.place(x=630, y=100)

# Image
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="images/upload_photo.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# Button
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000, y=370)
saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=1000, y=450)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000, y=610)

root.mainloop()